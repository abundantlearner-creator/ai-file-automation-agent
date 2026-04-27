import os
import shutil
import csv
from pypdf import PdfReader
from docx import Document
import openpyxl

source_folder = r"C:\Users\Owner\Downloads"
destination_base = r"C:\Users\Owner\Documents\Empress Command Center"


def read_pdf(file_path):
    try:
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages[:5]:
            text += page.extract_text() or ""
        return text.lower()
    except:
        return ""


def read_docx(file_path):
    try:
        doc = Document(file_path)
        return "\n".join([p.text for p in doc.paragraphs]).lower()
    except:
        return ""


def read_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        text = ""
        for sheet in wb.sheetnames[:2]:
            ws = wb[sheet]
            for row in ws.iter_rows(max_row=25, values_only=True):
                text += " ".join([str(cell) for cell in row if cell is not None]) + " "
        return text.lower()
    except:
        return ""


def read_csv_file(file_path):
    try:
        text = ""
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= 25:
                    break
                text += " ".join(row) + " "
        return text.lower()
    except:
        return ""


def get_file_text(file_path):
    name = os.path.basename(file_path).lower()

    if name.endswith(".pdf"):
        return read_pdf(file_path)
    elif name.endswith(".docx"):
        return read_docx(file_path)
    elif name.endswith((".xlsx", ".xls")):
        return read_excel(file_path)
    elif name.endswith(".csv"):
        return read_csv_file(file_path)
    elif name.endswith(".txt"):
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read(5000).lower()
        except:
            return ""
    else:
        return ""


def score_category(combined, keywords):
    score = 0
    for word, weight in keywords:
        if word in combined:
            score += weight
    return score


def suggest_folder(file_path):
    name = os.path.basename(file_path).lower()
    content = get_file_text(file_path)
    combined = name + " " + content

    # Media always wins by file type
    if name.endswith((
        ".png", ".jpg", ".jpeg", ".gif", ".webp",
        ".mp4", ".mov", ".avi", ".mkv",
        ".mp3", ".wav"
    )):
        return "Media"

    # Strong book/report/government PDF override
    if name.endswith(".pdf") and any(word in combined for word in [
        "table of contents",
        "chapter 1",
        "chapter 2",
        "introduction",
        "copyright",
        "all rights reserved",
        "isbn",
    ]):
        return "Books"

    categories = {
        "4IR": [
            ("4IR", 10),
            ("governance", 6),
            ("economic architecture", 8),
            ("personal governance", 8),
            ("capacity", 5),
            ("entity structure", 7),
            ("role map", 7),
            ("exposure", 5),
            ("mission engine", 7),
            ("holding company", 5),
            ("asset layer", 5),
            ("protection layer", 5),
        ],

        "Finance": [
            ("irs", 10),
            ("tax", 8),
            ("1040", 8),
            ("1099", 8),
            ("1098", 8),
            ("w2", 8),
            ("return", 5),
            ("transcript", 8),
            ("record of account", 10),
            ("trust", 8),
            ("trustee", 9),
            ("fiduciary", 9),
            ("form 56", 10),
            ("power of attorney", 10),
            ("poa", 8),
            ("certification of trust", 10),
            ("living trust", 10),
            ("consumer law", 10),
            ("fee schedule", 9),
            ("secured party", 8),
            ("creditor", 6),
        ],

        "Career": [
            ("resume", 10),
            ("professional summary", 9),
            ("work experience", 7),
            ("skills", 4),
            ("dear hiring manager", 10),
            ("cover letter", 10),
            ("job application", 8),
            ("linkedin", 7),
            ("hiring manager", 8),
            ("thank you for your time and consideration", 8),
            ("business operations", 5),
            ("application analyst", 5),
        ],

        "Trading": [
            ("futures", 10),
            ("options", 10),
            ("nq", 9),
            ("es", 7),
            ("mnq", 9),
            ("mes", 8),
            ("spy", 7),
            ("qqq", 7),
            ("iwm", 8),
            ("xsp", 7),
            ("spx", 7),
            ("delta", 8),
            ("theta", 8),
            ("open interest", 9),
            ("gamma", 8),
            ("atm", 7),
            ("itm", 7),
            ("otm", 7),
            ("0dte", 9),
            ("1dte", 8),
            ("2dte", 8),
            ("ticks", 8),
            ("points", 5),
            ("stop loss", 8),
            ("take profit", 8),
            ("vwap", 8),
            ("volume profile", 8),
            ("price action", 8),
            ("reclaim", 7),
            ("liquidity", 7),
            ("risk-to-reward", 9),
            ("bid", 4),
            ("ask", 4),
            ("strike", 7),
        ],

        "School": [
            ("math221", 10),
            ("hit227", 10),
            ("hit229", 10),
            ("hit277", 10),
            ("card205", 10),
            ("hit279", 10),
            ("ethc", 10),
            ("devry", 8),
            ("module", 8),
            ("assignment", 6),
            ("case study", 8),
            ("deficiency report", 9),
            ("probability", 9),
            ("probabilities", 9),
            ("binomial", 10),
            ("non-binomial", 10),
            ("distribution", 8),
            ("statistics", 9),
            ("variance", 8),
            ("standard deviation", 8),
            ("course project", 8),
            ("google analytics", 10),
            ("data analytics", 8),
            ("coursera", 8),
            ("dean's list", 10),
            ("academic achievement", 8),
        ],

        "Books": [
            ("book", 8),
            ("learn", 5),
            ("guide", 5),
            ("manual", 7),
            ("chapter", 5),
            ("table of contents", 10),
            ("oreilly", 10),
            ("pdfdrive", 10),
            ("learn python", 10),
            ("learn css", 10),
            ("learn c#", 10),
            ("mastering ethereum", 10),
            ("action plan", 8),
            ("policy", 5),
            ("report", 6),
        ],

        "Business": [
            ("amazon", 8),
            ("reseller", 8),
            ("net 30", 8),
            ("ecommerce", 8),
            ("virtual assistant", 8),
            ("invoice", 6),
            ("quote", 6),
            ("client", 4),
            ("order", 4),
            ("business", 4),
        ],

        "Vaccination_Background_Check": [
            ("vaccination", 10),
            ("immunization", 10),
            ("background check", 10),
            ("drug screen", 9),
            ("tb test", 9),
            ("clinical clearance", 8),
            ("practicum requirement", 8),
        ],
    }

    scores = {}

    for category, keywords in categories.items():
        scores[category] = score_category(combined, keywords)

    best_folder = max(scores, key=scores.get)
    best_score = scores[best_folder]

    # Prevent weak random matches
    if best_score < 8:
        return "Review_Needed"

    # Books only for PDF/ebook-style files
    if best_folder == "Books" and not name.endswith((".pdf", ".epub", ".mobi")):
        return "Review_Needed"

    return best_folder


files = [
    f for f in os.listdir(source_folder)
    if os.path.isfile(os.path.join(source_folder, f))
]

move_plan = []

for file in files:
    source_path = os.path.join(source_folder, file)
    folder = suggest_folder(source_path)
    destination_path = os.path.join(destination_base, folder, file)
    move_plan.append((source_path, destination_path, folder))


print("\n📦 SALLY SMART MODE PREVIEW:\n")

for src, dst, folder in move_plan[:25]:
    print(f"{os.path.basename(src)} → {folder}")

print(f"\nTotal files to review/move: {len(move_plan)}")

confirm = input("\nType 'yes' to move files: ")


if confirm.lower() == "yes":
    moved_count = 0
    skipped_count = 0

    for src, dst, folder in move_plan:
        try:
            os.makedirs(os.path.dirname(dst), exist_ok=True)

            if os.path.exists(src):
                shutil.move(src, dst)
                moved_count += 1
            else:
                skipped_count += 1

        except Exception as e:
            skipped_count += 1
            print(f"⚠️ Skipped: {os.path.basename(src)}")
            print(f"Reason: {e}\n")

    print(f"\n✅ Files moved: {moved_count}")
    print(f"⚠️ Files skipped: {skipped_count}")

else:
    print("\n❌ No files were moved.")